import json
import pytest
from pathlib import Path
import openpyxl

from backend.config import TEMPLATES_DIR, DB_PATH
from backend.models.schemas import DocumentTemplate, ExtractionStatus
from backend.services.batch_service import BatchService
from backend.services.database_service import DatabaseService
from backend.services.excel_service import ExcelService

@pytest.fixture
def workflow_setup(tmp_path):
    test_db = tmp_path / "workflow_test.db"
    db_service = DatabaseService(db_path=test_db)
    batch_service = BatchService(db_service=db_service)
    
    with open(TEMPLATES_DIR / "electricity_bill.json", "r", encoding="utf-8") as f:
        elec_tpl = DocumentTemplate(**json.load(f))
        
    with open(TEMPLATES_DIR / "invoice.json", "r", encoding="utf-8") as f:
        inv_tpl = DocumentTemplate(**json.load(f))
        
    return {
        "db": db_service,
        "batch": batch_service,
        "elec_tpl": elec_tpl,
        "inv_tpl": inv_tpl,
        "tmp_path": tmp_path
    }

def test_complete_end_to_end_workflow(workflow_setup):
    db = workflow_setup["db"]
    batch = workflow_setup["batch"]
    elec_tpl = workflow_setup["elec_tpl"]
    tmp_path = workflow_setup["tmp_path"]

    synthetic_dir = Path("sample_data/synthetic")
    assert synthetic_dir.exists()

    # 1. Gather all 18 electricity bill documents
    test_files = [synthetic_dir / f"bill_doc_{i:02d}.jpg" for i in range(1, 19)]
    for f in test_files:
        assert f.exists()

    # 2. Run batch processing
    summary = batch.process_batch(
        file_paths=test_files,
        template=elec_tpl,
        run_id="integration_run_01",
        max_workers=2
    )

    # 3. Confirm all files were processed without crashing
    assert summary.total_files == 18
    assert len(summary.documents) == 18

    # 4. Verify that the critically blurred file (bill_doc_18) is FAILED
    doc_18 = next((d for d in summary.documents if d.filename == "bill_doc_18.jpg"), None)
    assert doc_18 is not None
    assert doc_18.status == ExtractionStatus.FAILED
    assert doc_18.overall_confidence == 0.0

    # 5. Verify database persistence
    persisted_docs = db.list_documents(run_id="integration_run_01")
    assert len(persisted_docs) == 18

    # 6. Test manual correction on a document in Review queue
    review_doc = next((d for d in summary.documents if d.status == ExtractionStatus.REVIEW), None)
    assert review_doc is not None
    orig_doc_id = review_doc.document_id

    correction = db.save_manual_correction(
        doc_id=orig_doc_id,
        field_name="Current Reading",
        new_value=25000,
        notes="Operator confirmed manual correction"
    )
    assert correction["status"] == "success"

    # Confirm correction persisted in SQLite
    updated_doc = db.get_document(orig_doc_id)
    assert len(updated_doc["manual_corrections"]) >= 1
    assert updated_doc["manual_corrections"][0]["corrected_value"] == "25000"

    # 7. Export results to Excel workbook
    excel_path = ExcelService.export_results_to_excel(
        summary.documents, elec_tpl, "test_workflow_results.xlsx"
    )
    assert excel_path.exists()

    # 8. Programmatically inspect Excel workbook
    wb = openpyxl.load_workbook(str(excel_path))
    assert "Extracted Data" in wb.sheetnames
    assert "Needs Review" in wb.sheetnames
    assert "Failed" in wb.sheetnames
    assert "Processing Summary" in wb.sheetnames

    ws_good = wb["Extracted Data"]
    if ws_good.max_row >= 2:
        # Check Account Number column (col 3) is integer
        val = ws_good.cell(row=2, column=3).value
        assert isinstance(val, int)

    ws_failed = wb["Failed"]
    assert ws_failed.max_row >= 2 # At least bill_doc_18

    ws_summary = wb["Processing Summary"]
    assert ws_summary.cell(row=1, column=1).value == "BATCH PROCESSING SUMMARY"

    # 9. Verify CSV export
    csv_path = ExcelService.export_results_to_csv(
        summary.documents, elec_tpl, "test_workflow_results.csv"
    )
    assert csv_path.exists()
