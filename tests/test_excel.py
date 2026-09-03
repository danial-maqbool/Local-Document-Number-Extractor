import pytest
import openpyxl
from datetime import datetime
from backend.services.excel_service import ExcelService
from backend.models.schemas import (
    DocumentProcessResult, ExtractedFieldResult, QualityReport,
    QualityStatus, ExtractionStatus, DocumentTemplate, FieldDefinition, FieldType
)

@pytest.fixture
def sample_template():
    return DocumentTemplate(
        id="test_bill",
        name="Test Electricity Bill",
        fields=[
            FieldDefinition(name="Account No", type=FieldType.INTEGER),
            FieldDefinition(name="Current Reading", type=FieldType.INTEGER),
            FieldDefinition(name="Total Amount", type=FieldType.DECIMAL)
        ]
    )

@pytest.fixture
def sample_results(sample_template):
    qr_good = QualityReport(blur_score=100.0, brightness=120.0, contrast=50.0, width=800, height=1000, status=QualityStatus.GOOD)
    qr_fail = QualityReport(blur_score=10.0, brightness=20.0, contrast=10.0, width=200, height=200, status=QualityStatus.FAILED, issues=["Severe blur"])

    doc1 = DocumentProcessResult(
        document_id="doc_1", filename="img_01.jpg", file_hash="hash1", original_path="/path/1.jpg",
        template_id=sample_template.id, template_name=sample_template.name, quality=qr_good,
        fields={
            "Account No": ExtractedFieldResult(field_name="Account No", value=382947291, confidence=0.98, is_valid=True),
            "Current Reading": ExtractedFieldResult(field_name="Current Reading", value=14500, confidence=0.95, is_valid=True),
            "Total Amount": ExtractedFieldResult(field_name="Total Amount", value=3520.50, confidence=0.92, is_valid=True)
        },
        overall_confidence=0.95, status=ExtractionStatus.GOOD
    )

    doc2 = DocumentProcessResult(
        document_id="doc_2", filename="img_02.jpg", file_hash="hash2", original_path="/path/2.jpg",
        template_id=sample_template.id, template_name=sample_template.name, quality=qr_good,
        fields={
            "Account No": ExtractedFieldResult(field_name="Account No", value=918293812, confidence=0.50, is_valid=False, validation_notes=["Low confidence"]),
            "Current Reading": ExtractedFieldResult(field_name="Current Reading", value=8455, confidence=0.88, is_valid=True),
            "Total Amount": ExtractedFieldResult(field_name="Total Amount", value=2910.0, confidence=0.85, is_valid=True)
        },
        overall_confidence=0.62, status=ExtractionStatus.REVIEW, validation_errors=["Account No low confidence"]
    )

    doc3 = DocumentProcessResult(
        document_id="doc_3", filename="img_03.jpg", file_hash="hash3", original_path="/path/3.jpg",
        template_id=sample_template.id, template_name=sample_template.name, quality=qr_fail,
        fields={}, overall_confidence=0.0, status=ExtractionStatus.FAILED,
        validation_errors=["Critical blur detected"]
    )

    return [doc1, doc2, doc3]

def test_excel_export_structure(sample_template, sample_results):
    xlsx_path = ExcelService.export_results_to_excel(sample_results, sample_template, "test_output.xlsx")
    assert xlsx_path.exists()

    wb = openpyxl.load_workbook(str(xlsx_path))
    sheet_names = wb.sheetnames
    assert "Extracted Data" in sheet_names
    assert "Needs Review" in sheet_names
    assert "Failed" in sheet_names
    assert "Processing Summary" in sheet_names

    ws_good = wb["Extracted Data"]
    assert ws_good.max_row == 2 # 1 header + 1 record
    # Check numeric cell typing in Account No (col 3) and Current Reading (col 4)
    cell_acc = ws_good.cell(row=2, column=3)
    assert cell_acc.value == 382947291
    assert isinstance(cell_acc.value, int)

    cell_amount = ws_good.cell(row=2, column=5)
    assert cell_amount.value == 3520.50
    assert isinstance(cell_amount.value, float)

    ws_review = wb["Needs Review"]
    assert ws_review.max_row == 2 # 1 header + 1 record

    ws_failed = wb["Failed"]
    assert ws_failed.max_row == 2

    ws_sum = wb["Processing Summary"]
    assert ws_sum.cell(row=1, column=1).value == "BATCH PROCESSING SUMMARY"

def test_csv_export(sample_template, sample_results):
    csv_path = ExcelService.export_results_to_csv(sample_results, sample_template, "test_output.csv")
    assert csv_path.exists()
    content = csv_path.read_text(encoding="utf-8-sig")
    assert "Account No" in content
    assert "382947291" in content
