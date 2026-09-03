"""
Local Document Number Extractor - Multi-Sheet Excel & CSV Exporter
Generates professional .xlsx workbooks with typed numeric cells, formulas, styling, and multi-sheet summaries.
"""
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import logging

from backend.config import EXPORTS_DIR
from backend.models.schemas import DocumentProcessResult, DocumentTemplate, ExtractionStatus

logger = logging.getLogger("extractor.excel")

class ExcelService:

    @classmethod
    def export_results_to_excel(
        cls,
        results: List[DocumentProcessResult],
        template: DocumentTemplate,
        output_filename: Optional[str] = None
    ) -> Path:
        """
        Build and save a multi-sheet formatted Excel (.xlsx) workbook:
        Sheet 1: 'Extracted Data' (Good)
        Sheet 2: 'Needs Review' (Review)
        Sheet 3: 'Failed' (Failed)
        Sheet 4: 'Processing Summary'
        """
        if not output_filename:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            output_filename = f"extraction_results_{template.id}_{timestamp}.xlsx"

        output_path = EXPORTS_DIR / output_filename
        wb = openpyxl.Workbook()
        # Remove default empty sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Style definitions
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill_blue = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_fill_orange = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        header_fill_red = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        header_fill_gray = PatternFill(start_color="595959", end_color="595959", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        # Group results by status
        good_docs = [r for r in results if r.status == ExtractionStatus.GOOD]
        review_docs = [r for r in results if r.status == ExtractionStatus.REVIEW]
        failed_docs = [r for r in results if r.status == ExtractionStatus.FAILED]

        field_names = [f.name for f in template.fields]
        headers = ["Document ID", "Filename"] + field_names + [
            "Overall Confidence", "Status", "Cross-Field Check", "Validation Notes", "Processed At"
        ]

        # -------------------------------------------------------------
        # 1. Extracted Data Sheet
        # -------------------------------------------------------------
        ws_good = wb.create_sheet(title="Extracted Data")
        cls._populate_data_sheet(ws_good, headers, good_docs, field_names, header_font, header_fill_blue, thin_border)

        # -------------------------------------------------------------
        # 2. Needs Review Sheet
        # -------------------------------------------------------------
        ws_review = wb.create_sheet(title="Needs Review")
        cls._populate_data_sheet(ws_review, headers, review_docs, field_names, header_font, header_fill_orange, thin_border)

        # -------------------------------------------------------------
        # 3. Failed Sheet
        # -------------------------------------------------------------
        ws_failed = wb.create_sheet(title="Failed")
        failed_headers = ["Document ID", "Filename", "Original Path", "Failure Issues", "Status", "Processed At"]
        cls._populate_failed_sheet(ws_failed, failed_headers, failed_docs, header_font, header_fill_red, thin_border)

        # -------------------------------------------------------------
        # 4. Processing Summary Sheet
        # -------------------------------------------------------------
        ws_summary = wb.create_sheet(title="Processing Summary")
        cls._populate_summary_sheet(ws_summary, results, template, header_font, header_fill_gray, thin_border)

        wb.save(str(output_path))
        logger.info(f"Successfully generated Excel workbook: {output_path}")
        return output_path

    @classmethod
    def _populate_data_sheet(
        cls, ws, headers, docs, field_names, header_font, header_fill, border
    ):
        ws.append(headers)
        # Format Header row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for doc in docs:
            row_data = [doc.document_id, doc.filename]
            for fn in field_names:
                f_res = doc.fields.get(fn)
                val = f_res.value if f_res else None
                # Preserve numeric types
                if val is not None and isinstance(val, (int, float)):
                    row_data.append(val)
                elif val is not None:
                    row_data.append(str(val))
                else:
                    row_data.append("")

            row_data.append(round(doc.overall_confidence, 2))
            row_data.append(doc.status.value)
            row_data.append("PASSED" if doc.cross_field_validation_passed else "FAILED")
            row_data.append("; ".join(doc.validation_errors) if doc.validation_errors else "None")
            row_data.append(doc.processed_at.strftime("%Y-%m-%d %H:%M:%S"))
            ws.append(row_data)

        # Format Data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        # Freeze header & Auto filter
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

        cls._auto_fit_columns(ws)

    @classmethod
    def _populate_failed_sheet(cls, ws, headers, docs, header_font, header_fill, border):
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for doc in docs:
            issues = "; ".join(doc.quality.issues + doc.validation_errors) or "Unknown failure"
            ws.append([
                doc.document_id, doc.filename, doc.original_path,
                issues, doc.status.value, doc.processed_at.strftime("%Y-%m-%d %H:%M:%S")
            ])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="left")

        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
        cls._auto_fit_columns(ws)

    @classmethod
    def _populate_summary_sheet(cls, ws, results, template, header_font, header_fill, border):
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 30

        total = len(results)
        successful = sum(1 for r in results if r.status == ExtractionStatus.GOOD)
        review = sum(1 for r in results if r.status == ExtractionStatus.REVIEW)
        failed = sum(1 for r in results if r.status == ExtractionStatus.FAILED)
        non_failed_scores = [r.overall_confidence for r in results if r.status != ExtractionStatus.FAILED]
        avg_conf = round(sum(non_failed_scores) / len(non_failed_scores), 4) if non_failed_scores else 0.0

        title_cell = ws.cell(row=1, column=1, value="BATCH PROCESSING SUMMARY")
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="1F4E79")

        metrics = [
            ("Document Template", template.name),
            ("Total Documents Processed", total),
            ("Successful Documents (Good)", successful),
            ("Needs Review Documents", review),
            ("Failed Documents", failed),
            ("Success Rate", f"{(successful / total * 100):.1f}%" if total > 0 else "0%"),
            ("Average Confidence", f"{(avg_conf * 100):.1f}%"),
            ("Summary Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"))
        ]

        for idx, (label, val) in enumerate(metrics, start=3):
            lbl_cell = ws.cell(row=idx, column=1, value=label)
            lbl_cell.font = Font(name="Calibri", bold=True)
            lbl_cell.border = border

            val_cell = ws.cell(row=idx, column=2, value=val)
            val_cell.font = Font(name="Calibri")
            val_cell.border = border
            if isinstance(val, (int, float)):
                val_cell.alignment = Alignment(horizontal="right")

    @staticmethod
    def _auto_fit_columns(ws):
        """Auto-calculate and set reasonable column widths"""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    @classmethod
    def export_results_to_csv(
        cls,
        results: List[DocumentProcessResult],
        template: DocumentTemplate,
        output_filename: Optional[str] = None
    ) -> Path:
        """Export flat table to CSV format as secondary output"""
        if not output_filename:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            output_filename = f"extraction_results_{template.id}_{timestamp}.csv"

        output_path = EXPORTS_DIR / output_filename
        field_names = [f.name for f in template.fields]
        rows = []
        for doc in results:
            row = {
                "Document ID": doc.document_id,
                "Filename": doc.filename,
                "Status": doc.status.value,
                "Confidence": doc.overall_confidence
            }
            for fn in field_names:
                f_res = doc.fields.get(fn)
                row[fn] = f_res.value if f_res else None
            rows.append(row)

        df = pd.DataFrame(rows)
        df.to_csv(output_path, index=False, encoding="utf-8-sig")
        return output_path
